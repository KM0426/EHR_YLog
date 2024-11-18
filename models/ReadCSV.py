import csv,os
from lxml import etree
from openpyxl import load_workbook

def ReadCSV(paths):
    ret = []
    for path in paths:
        base_name, extension = os.path.splitext(path)
        if extension == '.csv':
            with open(path) as f:
                reader = csv.reader(f)
                for row in reader:
                    if len(row) > 0:
                        ret.append(row)
        if extension == '.xlsx':
            wb = load_workbook(path)
            sheet_name = wb.sheetnames[0]
            ws = wb[wb.sheetnames[0]]
            root = etree.Element("a"+sheet_name)

            for row in ws.iter_rows(values_only=True):
                xml_row = etree.SubElement(root, 'row')
                row_val = []
                for value in row:
                    str_val = str(value)
                    row_val.append(str(value))
                    # etree.SubElement(xml_row, 'cell').text = str(value)
                ret.append(row_val)

        # ret.append(etree.tostring(root, pretty_print=True).decode('utf-8'))

    return ret